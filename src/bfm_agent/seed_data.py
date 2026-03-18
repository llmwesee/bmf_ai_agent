from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from bfm_agent.config import get_settings
from bfm_agent.db import SCHEMA_VERSION, ensure_schema, reset_schema
from bfm_agent.models import (
    Account,
    AgentAction,
    AppConfig,
    BillingMilestone,
    InvoiceRecord,
    NotificationEvent,
    Project,
    RiskThreshold,
)


@dataclass(frozen=True)
class AccountSeed:
    name: str
    industry: str
    delivery_unit: str
    account_manager: str
    account_manager_email: str
    client_contact_name: str
    client_contact_email: str


@dataclass(frozen=True)
class ProjectSeed:
    account_name: str
    project_code: str
    project_name: str
    service_line: str
    billing_type: str
    billing_owner: str
    billing_owner_email: str
    contract_value: float
    revenue_plan_month: float
    revenue_recognized: float
    recognized_last_7_days: float
    recognized_last_30_days: float
    pending_pipeline: float
    forecast_bias: float
    last_revenue_update_days_ago: int
    billing_cycle_days: int
    start_days_ago: int
    end_days_ahead: int


@dataclass(frozen=True)
class MilestoneSeed:
    project_code: str
    milestone_name: str
    completion_days_ago: int
    billable_amount: float
    billed_amount: float
    invoice_generated: bool
    invoice_number: str | None
    invoice_days_ago: int | None
    account_manager_response: str
    status_note: str


@dataclass(frozen=True)
class InvoiceSeed:
    project_code: str
    invoice_number: str
    invoice_amount: float
    amount_received: float
    invoice_days_ago: int
    due_days_after_invoice: int
    collected_days_ago: int | None
    client_response_status: str
    status_note: str


@dataclass(frozen=True)
class ThresholdSeed:
    agent_key: str
    metric_key: str
    label: str
    unit: str
    medium_value: float
    high_value: float
    description: str


ACCOUNT_SEEDS: tuple[AccountSeed, ...] = (
    AccountSeed(
        "PepsiCo",
        "Consumer Goods",
        "Data & AI",
        "Priya Sharma",
        "priya.sharma@bfmdemo.local",
        "Daniel Cooper",
        "daniel.cooper@pepsico-demo.local",
    ),
    AccountSeed(
        "Comcast",
        "Telecom",
        "Digital Engineering",
        "Neha Iyer",
        "neha.iyer@bfmdemo.local",
        "Melissa Grant",
        "melissa.grant@comcast-demo.local",
    ),
    AccountSeed(
        "HSBC",
        "Banking",
        "Managed Services",
        "Rohit Desai",
        "rohit.desai@bfmdemo.local",
        "Liam Porter",
        "liam.porter@hsbc-demo.local",
    ),
    AccountSeed(
        "AstraZeneca",
        "Life Sciences",
        "Enterprise Platforms",
        "Kavya Nair",
        "kavya.nair@bfmdemo.local",
        "Sophia Reed",
        "sophia.reed@astrazeneca-demo.local",
    ),
    AccountSeed(
        "Cisco",
        "Networking",
        "Cyber Defense",
        "Sanjay Patel",
        "sanjay.patel@bfmdemo.local",
        "Victor Hill",
        "victor.hill@cisco-demo.local",
    ),
    AccountSeed(
        "Target",
        "Retail",
        "Commerce Platforms",
        "Ananya Rao",
        "ananya.rao@bfmdemo.local",
        "Grace Miller",
        "grace.miller@target-demo.local",
    ),
)


PROJECT_SEEDS: tuple[ProjectSeed, ...] = (
    ProjectSeed(
        "PepsiCo",
        "PRJ-4456",
        "Global Data Platform Rollout",
        "Data Modernization",
        "Milestone",
        "Aditya Bose",
        "aditya.bose@bfmdemo.local",
        3_600_000,
        1_400_000,
        1_100_000,
        35_000,
        140_000,
        30_000,
        0.95,
        6,
        7,
        320,
        420,
    ),
    ProjectSeed(
        "PepsiCo",
        "PRJ-4521",
        "Retail Insights Factory",
        "Analytics Engineering",
        "Time & Material",
        "Meera Jain",
        "meera.jain@bfmdemo.local",
        2_400_000,
        1_000_000,
        800_000,
        25_000,
        108_000,
        10_000,
        0.88,
        3,
        14,
        260,
        360,
    ),
    ProjectSeed(
        "Comcast",
        "PRJ-4588",
        "Customer App Modernization",
        "Digital Product Engineering",
        "Milestone",
        "Kiran Sethi",
        "kiran.sethi@bfmdemo.local",
        2_900_000,
        900_000,
        690_000,
        40_000,
        155_000,
        80_000,
        1.02,
        4,
        10,
        290,
        400,
    ),
    ProjectSeed(
        "Comcast",
        "PRJ-4620",
        "Network AIOps Pilot",
        "Platform Engineering",
        "Time & Material",
        "Rashmi Gupta",
        "rashmi.gupta@bfmdemo.local",
        2_200_000,
        750_000,
        580_000,
        28_000,
        118_000,
        45_000,
        0.98,
        5,
        14,
        180,
        330,
    ),
    ProjectSeed(
        "HSBC",
        "PRJ-4711",
        "Regulatory Data Platform",
        "Data Platforms",
        "Milestone",
        "Nikhil Verma",
        "nikhil.verma@bfmdemo.local",
        2_700_000,
        850_000,
        780_000,
        25_000,
        100_000,
        35_000,
        1.03,
        2,
        7,
        310,
        420,
    ),
    ProjectSeed(
        "HSBC",
        "PRJ-4730",
        "Treasury Operations Support",
        "Finance Ops",
        "Time & Material",
        "Asha Menon",
        "asha.menon@bfmdemo.local",
        2_050_000,
        650_000,
        590_000,
        18_000,
        88_000,
        20_000,
        1.0,
        2,
        14,
        240,
        390,
    ),
    ProjectSeed(
        "AstraZeneca",
        "PRJ-4870",
        "SAP Finance Transformation",
        "ERP Modernization",
        "Milestone",
        "Ritu Chopra",
        "ritu.chopra@bfmdemo.local",
        3_100_000,
        980_000,
        720_000,
        30_000,
        128_000,
        55_000,
        0.9,
        8,
        7,
        330,
        480,
    ),
    ProjectSeed(
        "AstraZeneca",
        "PRJ-4902",
        "Clinical Data Lake",
        "Data Platforms",
        "Time & Material",
        "Naveen Pillai",
        "naveen.pillai@bfmdemo.local",
        2_400_000,
        720_000,
        530_000,
        20_000,
        86_000,
        60_000,
        0.88,
        7,
        10,
        240,
        400,
    ),
    ProjectSeed(
        "Cisco",
        "PRJ-5031",
        "SOC Automation Program",
        "Cybersecurity",
        "Milestone",
        "Harsh Bhatia",
        "harsh.bhatia@bfmdemo.local",
        2_500_000,
        830_000,
        810_000,
        18_000,
        82_000,
        20_000,
        1.05,
        1,
        7,
        280,
        360,
    ),
    ProjectSeed(
        "Cisco",
        "PRJ-5077",
        "Secure Network Reliability",
        "Network Engineering",
        "Time & Material",
        "Tanya Singh",
        "tanya.singh@bfmdemo.local",
        2_050_000,
        620_000,
        600_000,
        15_000,
        70_000,
        10_000,
        1.02,
        2,
        14,
        220,
        330,
    ),
    ProjectSeed(
        "Target",
        "PRJ-5105",
        "Store Commerce Engine",
        "Commerce Engineering",
        "Milestone",
        "Varun Kapoor",
        "varun.kapoor@bfmdemo.local",
        2_300_000,
        780_000,
        745_000,
        20_000,
        92_000,
        15_000,
        1.0,
        1,
        7,
        250,
        360,
    ),
)


MILESTONE_SEEDS: tuple[MilestoneSeed, ...] = (
    MilestoneSeed("PRJ-4456", "Wave 3 billing package", 10, 220_000, 0, False, None, None, "Pending", "Revenue recognized but invoice trigger is pending."),
    MilestoneSeed("PRJ-4456", "Data migration sign-off", 7, 200_000, 0, False, None, None, "Pending", "Awaiting account manager approval to release invoice."),
    MilestoneSeed("PRJ-4521", "Sprint 18 billing lot", 4, 70_000, 50_000, True, "INV-4521-88", 2, "Pending", "Partially invoiced while change order is being approved."),
    MilestoneSeed("PRJ-4588", "Release train 7 sign-off", 8, 180_000, 120_000, True, "INV-4588-17", 3, "Pending", "Residual services effort is pending billing release."),
    MilestoneSeed("PRJ-4620", "Observability milestone", 6, 95_000, 40_000, True, "INV-4620-41", 4, "Pending", "Invoice created but not fully billed against completed effort."),
    MilestoneSeed("PRJ-4711", "Regulatory release billing", 2, 110_000, 110_000, True, "INV-4711-22", 1, "Completed", "Billing released on schedule."),
    MilestoneSeed("PRJ-4730", "Treasury operations monthly pack", 5, 85_000, 70_000, True, "INV-4730-11", 3, "Pending", "Variance pending customer acceptance."),
    MilestoneSeed("PRJ-4870", "ERP wave 3 milestone", 12, 210_000, 60_000, True, "INV-4870-18", 8, "Pending", "Large billing hold due to sign-off delay."),
    MilestoneSeed("PRJ-4902", "Clinical ingestion baseline", 9, 140_000, 35_000, True, "INV-4902-04", 7, "Pending", "Billing lag is impacting revenue realization."),
    MilestoneSeed("PRJ-5031", "Security automation release", 1, 75_000, 75_000, True, "INV-5031-31", 1, "Completed", "Current cycle billing completed."),
    MilestoneSeed("PRJ-5077", "Reliability squad monthly billing", 3, 65_000, 58_000, True, "INV-5077-09", 2, "Pending", "Minor pending effort remains to be billed."),
    MilestoneSeed("PRJ-5105", "Commerce engine go-live", 2, 90_000, 90_000, True, "INV-5105-02", 1, "Completed", "Milestone billed on time."),
)


INVOICE_SEEDS: tuple[InvoiceSeed, ...] = (
    InvoiceSeed("PRJ-4456", "INV-4410", 460_000, 460_000, 70, 30, 38, "Completed", "Payment collected within terms."),
    InvoiceSeed("PRJ-4456", "INV-4482", 220_000, 0, 62, 30, None, "Pending", "Client has not responded to the last reminder."),
    InvoiceSeed("PRJ-4521", "INV-4521-61", 730_000, 640_000, 28, 30, 6, "Promised", "Client committed to clear balance next week."),
    InvoiceSeed("PRJ-4588", "INV-4588-12", 540_000, 410_000, 45, 30, 10, "Pending", "Balance is still pending on approved invoice."),
    InvoiceSeed("PRJ-4620", "INV-4620-18", 515_000, 360_000, 36, 30, 8, "Promised", "Part-payment received, rest pending approval."),
    InvoiceSeed("PRJ-4711", "INV-4711-10", 720_000, 720_000, 24, 30, 4, "Completed", "Collected in full."),
    InvoiceSeed("PRJ-4730", "INV-4730-08", 535_000, 470_000, 33, 30, 9, "Promised", "Client confirmed payment processing."),
    InvoiceSeed("PRJ-4870", "INV-4870-01", 430_000, 180_000, 52, 30, 14, "Pending", "Collections follow-up is ongoing."),
    InvoiceSeed("PRJ-4902", "INV-4902-07", 325_000, 140_000, 39, 30, 11, "Pending", "Client disputes a milestone amount."),
    InvoiceSeed("PRJ-5031", "INV-5031-04", 790_000, 790_000, 21, 30, 5, "Completed", "Collected in full."),
    InvoiceSeed("PRJ-5077", "INV-5077-16", 585_000, 540_000, 26, 30, 7, "Promised", "Balance expected this cycle."),
    InvoiceSeed("PRJ-5105", "INV-5105-12", 740_000, 740_000, 18, 30, 4, "Completed", "Invoice closed."),
)


THRESHOLD_SEEDS: tuple[ThresholdSeed, ...] = (
    ThresholdSeed("revenue_realization", "forecast_shortfall_ratio", "Revenue shortfall ratio", "percent", 0.05, 0.12, "Gap between forecast and target used for revenue realization risk."),
    ThresholdSeed("revenue_realization", "revenue_delay_days", "Revenue update delay", "days", 4, 8, "Days since the last revenue recognition update."),
    ThresholdSeed("billing_trigger", "billing_delay_days", "Billing delay", "days", 4, 7, "Days between milestone completion and billing trigger."),
    ThresholdSeed("billing_trigger", "unbilled_amount", "Pending billable amount", "currency", 100_000, 250_000, "Unbilled milestone value that should already have triggered invoicing."),
    ThresholdSeed("unbilled_revenue", "days_unbilled", "Unbilled aging", "days", 5, 10, "How long revenue has remained recognized but not billed."),
    ThresholdSeed("unbilled_revenue", "unbilled_amount", "Unbilled revenue", "currency", 120_000, 300_000, "Revenue recognized but not yet invoiced."),
    ThresholdSeed("collection_monitoring", "overdue_days", "Overdue days", "days", 15, 30, "Days an invoice remains overdue."),
    ThresholdSeed("collection_monitoring", "outstanding_balance", "Outstanding balance", "currency", 100_000, 200_000, "Receivable value still open on the invoice."),
    ThresholdSeed("revenue_forecasting", "forecast_gap_ratio", "Forecast gap ratio", "percent", 0.05, 0.12, "Forecasted shortfall versus target."),
)

REQUIRED_SHEETS = {"Accounts", "Projects", "Billing Milestones", "Collections", "Thresholds"}


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _sheet_records(worksheet) -> list[dict[str, Any]]:
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value).strip() for value in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:] if any(value not in (None, "") for value in row)]


def generate_demo_workbook(path: Path, reference_date: date | None = None) -> int:
    today = reference_date or date.today()
    workbook = Workbook()
    workbook.remove(workbook.active)

    account_totals: dict[str, dict[str, float]] = {}
    for seed in PROJECT_SEEDS:
        account_totals.setdefault(seed.account_name, {"monthly_target": 0.0, "contract_value": 0.0})
        account_totals[seed.account_name]["monthly_target"] += seed.revenue_plan_month
        account_totals[seed.account_name]["contract_value"] += seed.contract_value

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Sheet", "Purpose"])
    instructions.append(["Accounts", "Account owners, delivery units, client contacts, and monthly targets"])
    instructions.append(["Projects", "Raw revenue, trend, forecast, and billing ownership inputs"])
    instructions.append(["Billing Milestones", "Milestone completion and billing trigger records"])
    instructions.append(["Collections", "Invoice and collection tracking records"])
    instructions.append(["Thresholds", "Agent-specific dynamic risk thresholds editable from the UI"])

    accounts_sheet = workbook.create_sheet("Accounts")
    accounts_sheet.append(
        [
            "Account Name",
            "Industry",
            "Delivery Unit",
            "Account Manager",
            "Account Manager Email",
            "Client Contact Name",
            "Client Contact Email",
            "Account Monthly Target",
            "Contract Value",
        ]
    )
    for account in ACCOUNT_SEEDS:
        totals = account_totals[account.name]
        accounts_sheet.append(
            [
                account.name,
                account.industry,
                account.delivery_unit,
                account.account_manager,
                account.account_manager_email,
                account.client_contact_name,
                account.client_contact_email,
                totals["monthly_target"],
                totals["contract_value"],
            ]
        )

    projects_sheet = workbook.create_sheet("Projects")
    projects_sheet.append(
        [
            "Account Name",
            "Project Code",
            "Project Name",
            "Service Line",
            "Billing Type",
            "Billing Owner",
            "Billing Owner Email",
            "Contract Value",
            "Revenue Plan (Month)",
            "Revenue Plan (Quarter)",
            "Revenue Recognized",
            "Recognized Last 7 Days",
            "Recognized Last 30 Days",
            "Pending Pipeline",
            "Forecast Bias",
            "Last Revenue Update",
            "Billing Cycle Days",
            "Start Date",
            "End Date",
        ]
    )
    for seed in PROJECT_SEEDS:
        projects_sheet.append(
            [
                seed.account_name,
                seed.project_code,
                seed.project_name,
                seed.service_line,
                seed.billing_type,
                seed.billing_owner,
                seed.billing_owner_email,
                seed.contract_value,
                seed.revenue_plan_month,
                seed.revenue_plan_month * 3,
                seed.revenue_recognized,
                seed.recognized_last_7_days,
                seed.recognized_last_30_days,
                seed.pending_pipeline,
                seed.forecast_bias,
                (today - timedelta(days=seed.last_revenue_update_days_ago)).isoformat(),
                seed.billing_cycle_days,
                (today - timedelta(days=seed.start_days_ago)).isoformat(),
                (today + timedelta(days=seed.end_days_ahead)).isoformat(),
            ]
        )

    billing_sheet = workbook.create_sheet("Billing Milestones")
    billing_sheet.append(
        [
            "Project Code",
            "Milestone Name",
            "Completion Date",
            "Billable Amount",
            "Billed Amount",
            "Invoice Generated",
            "Invoice Number",
            "Invoice Date",
            "Account Manager Response",
            "Status Note",
        ]
    )
    for seed in MILESTONE_SEEDS:
        invoice_date = today - timedelta(days=seed.invoice_days_ago) if seed.invoice_days_ago is not None else ""
        billing_sheet.append(
            [
                seed.project_code,
                seed.milestone_name,
                (today - timedelta(days=seed.completion_days_ago)).isoformat(),
                seed.billable_amount,
                seed.billed_amount,
                "Yes" if seed.invoice_generated else "No",
                seed.invoice_number or "",
                invoice_date.isoformat() if invoice_date else "",
                seed.account_manager_response,
                seed.status_note,
            ]
        )

    collections_sheet = workbook.create_sheet("Collections")
    collections_sheet.append(
        [
            "Project Code",
            "Invoice Number",
            "Invoice Amount",
            "Amount Received",
            "Invoice Date",
            "Due Date",
            "Collected Date",
            "Client Response Status",
            "Status Note",
        ]
    )
    for seed in INVOICE_SEEDS:
        invoice_date = today - timedelta(days=seed.invoice_days_ago)
        due_date = invoice_date + timedelta(days=seed.due_days_after_invoice)
        collected_date = today - timedelta(days=seed.collected_days_ago) if seed.collected_days_ago is not None else ""
        collections_sheet.append(
            [
                seed.project_code,
                seed.invoice_number,
                seed.invoice_amount,
                seed.amount_received,
                invoice_date.isoformat(),
                due_date.isoformat(),
                collected_date.isoformat() if collected_date else "",
                seed.client_response_status,
                seed.status_note,
            ]
        )

    thresholds_sheet = workbook.create_sheet("Thresholds")
    thresholds_sheet.append(
        [
            "Agent Key",
            "Metric Key",
            "Label",
            "Unit",
            "Medium Value",
            "High Value",
            "Description",
        ]
    )
    for threshold in THRESHOLD_SEEDS:
        thresholds_sheet.append(
            [
                threshold.agent_key,
                threshold.metric_key,
                threshold.label,
                threshold.unit,
                threshold.medium_value,
                threshold.high_value,
                threshold.description,
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return len(PROJECT_SEEDS)


def _workbook_is_current(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception:
        return False
    try:
        return REQUIRED_SHEETS.issubset(set(workbook.sheetnames))
    finally:
        workbook.close()


def import_workbook_into_db(workbook_path: Path, session: Session) -> int:
    workbook = load_workbook(workbook_path, data_only=True)
    accounts = _sheet_records(workbook["Accounts"])
    projects = _sheet_records(workbook["Projects"])
    milestones = _sheet_records(workbook["Billing Milestones"])
    collections = _sheet_records(workbook["Collections"])
    thresholds = _sheet_records(workbook["Thresholds"])

    session.execute(delete(NotificationEvent))
    session.execute(delete(AgentAction))
    session.execute(delete(RiskThreshold))
    session.execute(delete(InvoiceRecord))
    session.execute(delete(BillingMilestone))
    session.execute(delete(Project))
    session.execute(delete(Account))
    session.flush()

    account_lookup: dict[str, Account] = {}
    for row in accounts:
        account = Account(
            name=str(row["Account Name"]),
            industry=str(row["Industry"]),
            delivery_unit=str(row["Delivery Unit"]),
            account_manager=str(row["Account Manager"]),
            account_manager_email=str(row["Account Manager Email"]),
            client_contact_name=str(row["Client Contact Name"]),
            client_contact_email=str(row["Client Contact Email"]),
            monthly_target=float(row["Account Monthly Target"]),
            contract_value=float(row["Contract Value"]),
        )
        session.add(account)
        session.flush()
        account_lookup[account.name] = account

    project_lookup: dict[str, Project] = {}
    for row in projects:
        project = Project(
            account_id=account_lookup[str(row["Account Name"])].id,
            code=str(row["Project Code"]),
            name=str(row["Project Name"]),
            service_line=str(row["Service Line"]),
            billing_type=str(row["Billing Type"]),
            billing_owner=str(row["Billing Owner"]),
            billing_owner_email=str(row["Billing Owner Email"]),
            contract_value=float(row["Contract Value"]),
            revenue_plan_month=float(row["Revenue Plan (Month)"]),
            revenue_plan_quarter=float(row["Revenue Plan (Quarter)"]),
            revenue_recognized=float(row["Revenue Recognized"]),
            recognized_last_7_days=float(row["Recognized Last 7 Days"]),
            recognized_last_30_days=float(row["Recognized Last 30 Days"]),
            pending_pipeline=float(row["Pending Pipeline"]),
            forecast_bias=float(row["Forecast Bias"]),
            last_revenue_update=_parse_date(row["Last Revenue Update"]),
            billing_cycle_days=int(row["Billing Cycle Days"]),
            start_date=_parse_date(row["Start Date"]),
            end_date=_parse_date(row["End Date"]),
        )
        session.add(project)
        session.flush()
        project_lookup[project.code] = project

    for row in milestones:
        invoice_generated = str(row["Invoice Generated"]).strip().lower() == "yes"
        invoice_date = row["Invoice Date"]
        project = project_lookup[str(row["Project Code"])]
        session.add(
            BillingMilestone(
                project_id=project.id,
                milestone_name=str(row["Milestone Name"]),
                completion_date=_parse_date(row["Completion Date"]),
                billable_amount=float(row["Billable Amount"]),
                billed_amount=float(row["Billed Amount"]),
                invoice_generated=invoice_generated,
                invoice_number=str(row["Invoice Number"]) if row["Invoice Number"] not in (None, "") else None,
                invoice_date=_parse_date(invoice_date) if invoice_date not in (None, "") else None,
                billing_owner=project.billing_owner,
                billing_owner_email=project.billing_owner_email,
                account_manager_response=str(row["Account Manager Response"]),
                status_note=str(row["Status Note"]),
            )
        )

    for row in collections:
        project = project_lookup[str(row["Project Code"])]
        account = project.account
        collected_date = row["Collected Date"]
        session.add(
            InvoiceRecord(
                project_id=project.id,
                invoice_number=str(row["Invoice Number"]),
                invoice_amount=float(row["Invoice Amount"]),
                amount_received=float(row["Amount Received"]),
                invoice_date=_parse_date(row["Invoice Date"]),
                due_date=_parse_date(row["Due Date"]),
                collected_date=_parse_date(collected_date) if collected_date not in (None, "") else None,
                client_contact_name=account.client_contact_name,
                client_contact_email=account.client_contact_email,
                collection_owner=project.billing_owner,
                collection_owner_email=project.billing_owner_email,
                client_response_status=str(row["Client Response Status"]),
                status_note=str(row["Status Note"]),
            )
        )

    for row in thresholds:
        session.add(
            RiskThreshold(
                agent_key=str(row["Agent Key"]),
                metric_key=str(row["Metric Key"]),
                label=str(row["Label"]),
                unit=str(row["Unit"]),
                medium_value=float(row["Medium Value"]),
                high_value=float(row["High Value"]),
                description=str(row["Description"]),
            )
        )

    session.merge(AppConfig(key="schema_version", value=SCHEMA_VERSION))
    session.flush()
    return len(projects)


def bootstrap_demo_assets(force_regenerate: bool = False) -> int:
    settings = get_settings()
    if force_regenerate or not _workbook_is_current(settings.workbook_file):
        generate_demo_workbook(settings.workbook_file)

    if force_regenerate:
        reset_schema()
    else:
        ensure_schema()

    from bfm_agent.db import session_scope

    with session_scope() as session:
        has_accounts = session.scalar(select(Account.id).limit(1))
        if force_regenerate or not has_accounts:
            return import_workbook_into_db(settings.workbook_file, session)
        return len(PROJECT_SEEDS)
